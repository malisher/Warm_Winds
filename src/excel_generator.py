import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from copy import deepcopy

from src.utils import pixels_to_width_units, get_image_size_with_aspect_ratio
import src.config as config

def generate_excel(images_changed_data, data_folder, cover_all):
    wb = openpyxl.Workbook()
    ws = wb.active

    result = []
    max_points = 0
    copied_data = deepcopy(images_changed_data)

    for image_data in copied_data:
        label_groups = {}
        for changed_point in image_data['children']['box']:
            if changed_point['label'] == 'ignore':
                continue
            else:
            # if cover_all is True or changed_point['label'] != 'ignore':
                label = changed_point['label']
                if label not in label_groups:
                    label_groups[label] = []
                label_groups[label].append(changed_point)

        for label, points in label_groups.items():
            image_file_paths = []
            for point in points:
                local_file_path = f"{data_folder}/cropped/{point['cropped_name']}"
                image_file_paths.append(local_file_path)

            column = [label, None, None, image_data['attributes']['id'], image_data['attributes']['name'], len(points)] + image_file_paths

            if len(points) > max_points:
                max_points = len(points)

            result.append(column)

    result.sort(key=lambda x: x[0])

    headers = config.headers + [f'фото{x}' for x in range(1, max_points + 1)]
    ws.append(headers)

    max_image_widths = [0] * max_points

    for row, data in enumerate(result, start=2):
        max_height_in_row = 0
        for col, cell_value in enumerate(data, start=1):
            if isinstance(cell_value, str) and cell_value.startswith(f"{data_folder}/cropped/"):
                img_width, img_height = get_image_size_with_aspect_ratio(cell_value, config.img_max_width)
                img = Image(cell_value)
                img.width, img.height = img_width, img_height
                max_height_in_row = max(max_height_in_row, img.height)
                ws.add_image(img, get_column_letter(col) + str(row))

                image_col_index = col - config.headers_count
                if 0 <= image_col_index < len(max_image_widths):
                    max_image_widths[image_col_index] = max(max_image_widths[image_col_index], img_width)
            else:
                ws.cell(row=row, column=col, value=cell_value)

        if max_height_in_row > 0:
            ws.row_dimensions[row].height = max_height_in_row * config.row_height_coef

    for i, max_width in enumerate(max_image_widths, start=config.headers_count):
        if max_width > 0:
            ws.column_dimensions[get_column_letter(i)].width = pixels_to_width_units(max_width)

    ws.column_dimensions['A'].width = config.column_a_width
    ws.column_dimensions['E'].width = config.column_e_width

    output_path = f'{data_folder}/output.xlsx'
    wb.save(output_path)
